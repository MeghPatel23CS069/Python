import openpyxl as op
import pandas as pd

Events=["select","Hackathon","Spoural","Cultural fest","Tech fest","Cognizance"]
path="file.xlsx"


# Save the excel file
def save_name(name,dept,contact,event,attendance) :
    try :
        workbook = op.load_workbook(path)
        sheet = workbook.active
        sheet.append([name,dept,contact,event,attendance])
    except FileNotFoundError :
        workbook = op.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Department", "Contact", "Event", "Attendance"])
        sheet.append([name,dept,contact,event,attendance])

    workbook.save(path)

# Fetch participants for a specific event
def get_participants_for_event(event):
    workbook = op.load_workbook(path)
    sheet = workbook.active
    participants = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[3] == event:
            participants.append(row[0])
    return participants


# Get the details of participants by name
def get_details_by_name(name) :
    workbook=op.load_workbook(path)
    sheet=workbook.active

    names=[]

    for row in sheet.iter_rows(min_row=2,values_only=True) :
        if row[0]==name :
            names.append(row)
    return names